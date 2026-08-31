from __future__ import annotations

import json
import os
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Page, sync_playwright

REPORT_URL = "https://sinapp.dgehm.gob.sv/drhm/estadisticas.aspx?uid=2"
DATA_DIR = Path("data")
DOWNLOAD_DIR = Path("downloads")


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(r"[^a-z0-9]+", " ", text.encode("ascii", "ignore").decode().lower()).strip()


def download_report(page: Page) -> Path:
    try:
        # El visor SSRS de DGEHM suele dejar recursos secundarios pendientes.
        # Esperar "load" hace que Playwright venza aun cuando el reporte ya está visible.
        page.goto(REPORT_URL, wait_until="domcontentloaded", timeout=180_000)
        page.wait_for_timeout(15_000)
    except Exception:
        Path("diagnostic.html").write_text(page.content(), encoding="utf-8")
        page.screenshot(path="diagnostic.png", full_page=True)
        raise

    contexts = [page, *page.frames]
    export_selectors = [
        "[title*='Export']",
        "[aria-label*='Export']",
        "img[alt*='Export']",
        "#ReportViewer1_ctl05_ctl04_ctl00_ButtonImg",
        "#ReportViewer1_ctl05_ctl04_ctl00",
    ]

    for context in contexts:
        for selector in export_selectors:
            button = context.locator(selector).first
            try:
                if button.is_visible(timeout=1_000):
                    button.click()
                    page.wait_for_timeout(800)
                    for candidate in ["Excel", "Excel 2007", "Archivo de Excel"]:
                        item = context.get_by_text(candidate, exact=False).last
                        try:
                            if item.is_visible(timeout=800):
                                with page.expect_download(timeout=120_000) as info:
                                    item.click()
                                download = info.value
                                DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
                                target = DOWNLOAD_DIR / "dgehm_latest.xlsx"
                                download.save_as(target)
                                return target
                        except Exception:
                            continue
            except Exception:
                continue

    Path("diagnostic.html").write_text(page.content(), encoding="utf-8")
    page.screenshot(path="diagnostic.png", full_page=True)
    raise RuntimeError("No se encontró la opción de exportación Excel del reporte DGEHM")


ALIASES = {
    "station": ["nombre estacion", "estacion de servicio", "estacion", "nombre comercial"],
    "brand": ["bandera", "marca", "compania", "distribuidor"],
    "department": ["departamento"],
    "municipality": ["municipio", "distrito"],
    "address": ["direccion", "ubicacion"],
    "product": ["producto", "combustible"],
    "price": ["precio", "precio venta"],
    "regular": ["gasolina regular", "regular"],
    "superior": ["gasolina especial", "gasolina superior", "especial", "superior", "premium"],
    "diesel": ["diesel bajo en azufre", "diesel"],
    "updated": ["ultima actualizacion", "fecha actualizacion", "actualizacion", "fecha"],
}


def find_header(sheet) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(sheet.max_row, 40) + 1):
        values = [normalized(sheet.cell(row_number, col).value) for col in range(1, sheet.max_column + 1)]
        found: dict[str, int] = {}
        for field, aliases in ALIASES.items():
            for col, value in enumerate(values, 1):
                if any(alias == value or alias in value for alias in aliases):
                    found[field] = col
                    break
        if "station" in found and ("price" in found or len({"regular", "superior", "diesel"} & found.keys()) >= 2):
            return row_number, found
    raise ValueError(f"No se reconocieron los encabezados en la hoja {sheet.title}")


def number(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(re.sub(r"[^0-9.]", "", str(value))), 5)
    except ValueError:
        return None


def parse_report(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = max(workbook.worksheets, key=lambda item: item.max_row * item.max_column)
    header_row, columns = find_header(sheet)
    captured_at = datetime.now(timezone.utc).isoformat()
    combined: dict[tuple[str, str], dict] = {}

    for row in range(header_row + 1, sheet.max_row + 1):
        get = lambda field: sheet.cell(row, columns[field]).value if field in columns else None
        station = str(get("station") or "").strip()
        if not station:
            continue
        record = {
            "station": station,
            "brand": str(get("brand") or "Bandera blanca").strip(),
            "department": str(get("department") or "").strip(),
            "municipality": str(get("municipality") or "").strip(),
            "address": str(get("address") or "").strip(),
            "regular": number(get("regular")),
            "superior": number(get("superior")),
            "diesel": number(get("diesel")),
            "capturedAt": captured_at,
            "sourceUrl": REPORT_URL,
        }
        if "product" in columns and "price" in columns:
            product = normalized(get("product"))
            price = number(get("price"))
            key = normalized(station), normalized(record["address"])
            target = combined.setdefault(key, record)
            if "regular" in product:
                target["regular"] = price
            elif any(item in product for item in ("premium", "superior", "especial")):
                target["superior"] = price
            elif "diesel" in product:
                target["diesel"] = price
        else:
            combined[(normalized(station), normalized(record["address"]))] = record

    rows = [value for value in combined.values() if any(value[item] is not None for item in ("regular", "superior", "diesel"))]
    if len(rows) < 100:
        raise ValueError(f"La extracción produjo solo {len(rows)} estaciones; no se publicará una carga incompleta")
    return sorted(rows, key=lambda item: (item["department"], item["municipality"], item["station"]))


def save(rows: list[dict], source: Path) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    history = DATA_DIR / "history"
    history.mkdir(parents=True, exist_ok=True)
    today = datetime.now(timezone.utc).date().isoformat()
    payload = {"source": REPORT_URL, "capturedAt": datetime.now(timezone.utc).isoformat(), "count": len(rows), "prices": rows}
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    (DATA_DIR / "latest.json").write_text(text, encoding="utf-8")
    (history / f"{today}.json").write_text(text, encoding="utf-8")
    shutil.copy2(source, DATA_DIR / "latest.xlsx")


def main() -> None:
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True, ignore_https_errors=os.getenv("DGEHM_IGNORE_HTTPS_ERRORS", "true").lower() == "true")
        page = context.new_page()
        try:
            report = download_report(page)
            save(parse_report(report), report)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
